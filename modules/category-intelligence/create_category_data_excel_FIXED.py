#!/usr/bin/env python3
"""
Create 04_CATEGORY_DATA.XLSX with CORRECT field mappings.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Define data directory
DATA_DIR = Path(__file__).parent / "data" / "retailers"

print("=" * 70)
print("CREATING 04_CATEGORY_DATA.XLSX (CORRECTED VERSION)")
print("=" * 70)
print()

# Load all retailer data
all_products = []

# Amazon
print("Loading Amazon products...")
try:
    with open(DATA_DIR / "amazon_products.json", 'r', encoding='utf-8') as f:
        amazon_data = json.load(f)
        for product in amazon_data:
            # Extract category/subcategory from taxonomy_path
            taxonomy = product.get('taxonomy_path', [])
            category = taxonomy[1] if len(taxonomy) > 1 else ''
            subcategory = taxonomy[2] if len(taxonomy) > 2 else ''

            # Get specs from attributes for description
            attributes = product.get('attributes', {})
            specs_list = []
            if isinstance(attributes, dict):
                for key, value in attributes.items():
                    if key != 'categories' and isinstance(value, (str, int, float)):
                        specs_list.append(f"{key}: {value}")
            description = "; ".join(specs_list) if specs_list else ""

            all_products.append({
                'Retailer': 'Amazon',
                'Product ID': product.get('sku', ''),
                'Product Name': product.get('name', ''),  # CORRECTED: was 'title'
                'Brand': product.get('brand', ''),
                'Category': category,
                'Subcategory': subcategory,
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('reviews', 0),  # CORRECTED: was 'total_reviews'
                'Description': description,
                'Product URL': product.get('url', ''),
                'Image URL': '',  # Not scraped for Amazon
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Amazon'])} Amazon products")
except Exception as e:
    print(f"  ✗ Error loading Amazon: {e}")

# Walmart
print("Loading Walmart products...")
try:
    with open(DATA_DIR / "walmart_products.json", 'r', encoding='utf-8') as f:
        walmart_data = json.load(f)
        for product in walmart_data:
            # Extract category/subcategory from taxonomy_path
            taxonomy = product.get('taxonomy_path', [])
            category = taxonomy[1] if len(taxonomy) > 1 else ''
            subcategory = taxonomy[2] if len(taxonomy) > 2 else ''

            # Get specs from attributes for description
            attributes = product.get('attributes', {})
            specs = attributes.get('specs', [])
            specs_list = []
            if isinstance(specs, list):
                for spec in specs:
                    if isinstance(spec, dict):
                        name = spec.get('name', '')
                        value = spec.get('value', '')
                        if name and value:
                            specs_list.append(f"{name}: {value}")
            description = "; ".join(specs_list) if specs_list else ""

            all_products.append({
                'Retailer': 'Walmart',
                'Product ID': product.get('sku', ''),
                'Product Name': product.get('name', ''),  # CORRECTED
                'Brand': product.get('brand', ''),
                'Category': category,
                'Subcategory': subcategory,
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('reviews', 0),  # CORRECTED
                'Description': description,
                'Product URL': product.get('url', ''),
                'Image URL': '',  # Not scraped for Walmart
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Walmart'])} Walmart products")
except Exception as e:
    print(f"  ✗ Error loading Walmart: {e}")

# Home Depot
print("Loading Home Depot products...")
try:
    with open(DATA_DIR / "homedepot_products.json", 'r', encoding='utf-8') as f:
        homedepot_data = json.load(f)
        for product in homedepot_data:
            # Extract category/subcategory from taxonomy_path
            taxonomy = product.get('taxonomy_path', [])
            category = taxonomy[1] if len(taxonomy) > 1 else ''
            subcategory = taxonomy[2] if len(taxonomy) > 2 else ''

            # Get attributes for description
            attributes = product.get('attributes', {})
            specs_list = []
            if isinstance(attributes, dict):
                for key, value in attributes.items():
                    if isinstance(value, (str, int, float)):
                        specs_list.append(f"{key}: {value}")
            description = "; ".join(specs_list) if specs_list else ""

            all_products.append({
                'Retailer': 'Home Depot',
                'Product ID': product.get('sku', ''),
                'Product Name': product.get('name', ''),  # CORRECTED
                'Brand': product.get('brand', ''),
                'Category': category,
                'Subcategory': subcategory,
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('reviews', 0),  # CORRECTED
                'Description': description,
                'Product URL': product.get('url', ''),
                'Image URL': '',  # Not scraped for Home Depot
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Home Depot'])} Home Depot products")
except Exception as e:
    print(f"  ✗ Error loading Home Depot: {e}")

# Target
print("Loading Target products...")
try:
    with open(DATA_DIR / "target_products.json", 'r', encoding='utf-8') as f:
        target_data = json.load(f)
        for product in target_data:
            # Target uses different field names
            taxonomy = product.get('taxonomy', '')
            category = taxonomy if taxonomy else ''

            # Extract rating from rating_text (e.g., "4.5 out of 5")
            rating_text = product.get('rating_text', '')
            rating = None
            if rating_text:
                try:
                    rating = float(rating_text.split()[0])
                except:
                    pass

            all_products.append({
                'Retailer': 'Target',
                'Product ID': '',  # Target data doesn't have SKU
                'Product Name': product.get('title', ''),  # CORRECTED: Target uses 'title'
                'Brand': product.get('brand', ''),
                'Category': category,
                'Subcategory': '',
                'Price': product.get('price'),
                'Rating': rating,
                'Review Count': 0,  # Not in Target data
                'Description': '',
                'Product URL': product.get('url', ''),
                'Image URL': '',  # Not scraped for Target
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Target'])} Target products")
except Exception as e:
    print(f"  ✗ Error loading Target: {e}")

# Etsy
print("Loading Etsy products...")
try:
    with open(DATA_DIR / "etsy_products.json", 'r', encoding='utf-8') as f:
        etsy_data = json.load(f)
        for product in etsy_data:
            # Etsy uses different field names
            all_products.append({
                'Retailer': 'Etsy',
                'Product ID': '',  # Etsy data doesn't have clear SKU
                'Product Name': product.get('title', ''),  # CORRECTED: Etsy uses 'title'
                'Brand': product.get('shop', ''),  # CORRECTED: Etsy uses 'shop' not 'brand'
                'Category': '',
                'Subcategory': '',
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': 0,  # Not in Etsy data
                'Description': product.get('badge', ''),  # Use badge as description
                'Product URL': product.get('url', ''),
                'Image URL': product.get('image', ''),  # Etsy has image URLs
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Etsy'])} Etsy products")
except Exception as e:
    print(f"  ✗ Error loading Etsy: {e}")

print()
print(f"Total products loaded: {len(all_products)}")
print()

# Create DataFrame
df = pd.DataFrame(all_products)

# Reorder columns
column_order = [
    'Retailer',
    'Product ID',
    'Product Name',
    'Brand',
    'Category',
    'Subcategory',
    'Price',
    'Rating',
    'Review Count',
    'Description',
    'Product URL',
    'Image URL'
]

df = df[column_order]

# Format data types
df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
df['Rating'] = pd.to_numeric(df['Rating'], errors='coerce')
df['Review Count'] = pd.to_numeric(df['Review Count'], errors='coerce').fillna(0).astype(int)

# Sort by Retailer and then by Product Name
df = df.sort_values(['Retailer', 'Product Name'])

# Save to Excel
output_file = Path(__file__).parent / "04_CATEGORY_DATA.xlsx"
print(f"Saving to Excel: {output_file.name}")
df.to_excel(output_file, index=False, sheet_name='Product Data')

# Format Excel file
print("Applying Excel formatting...")
wb = load_workbook(output_file)
ws = wb['Product Data']

# Header formatting
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
column_widths = {
    'A': 12,  # Retailer
    'B': 15,  # Product ID
    'C': 50,  # Product Name
    'D': 20,  # Brand
    'E': 25,  # Category
    'F': 25,  # Subcategory
    'G': 10,  # Price
    'H': 8,   # Rating
    'I': 12,  # Review Count
    'J': 60,  # Description
    'K': 50,  # Product URL
    'L': 50   # Image URL
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Format price columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row:
        if cell.value:
            cell.number_format = '$#,##0.00'

# Format rating columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
    for cell in row:
        if cell.value:
            cell.number_format = '0.0'

# Format review count columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
    for cell in row:
        if cell.value:
            cell.number_format = '#,##0'

# Freeze header row
ws.freeze_panes = 'A2'

# Add filters
ws.auto_filter.ref = ws.dimensions

# Save formatted workbook
wb.save(output_file)

print()
print("=" * 70)
print("EXCEL FILE CREATED SUCCESSFULLY")
print("=" * 70)
print(f"File: {output_file.name}")
print(f"Total products: {len(df):,}")
print()
print("Breakdown by retailer:")
for retailer in df['Retailer'].unique():
    count = len(df[df['Retailer'] == retailer])
    print(f"  {retailer}: {count:,} products")
print()
print("NOTE: Image URLs only available for Etsy. Descriptions compiled from")
print("      available product specifications where possible.")
print()
