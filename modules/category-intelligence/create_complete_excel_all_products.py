#!/usr/bin/env python3
"""
Create final Excel with ALL scraped products (no filtering).
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re

print("="*70)
print("CREATING COMPLETE DELIVERABLE WITH ALL PRODUCTS")
print("="*70)
print()

# Heuristic extraction functions
def extract_material(text):
    text_lower = text.lower()
    if any(word in text_lower for word in ['steel', 'stainless steel', 'metal']):
        return 'metal'
    elif 'aluminum' in text_lower or 'aluminium' in text_lower:
        return 'aluminum'
    elif 'plastic' in text_lower or 'resin' in text_lower:
        return 'plastic'
    elif 'wood' in text_lower or 'wooden' in text_lower:
        return 'wood'
    return 'unknown'

def extract_color(text):
    text_lower = text.lower()
    colors = ['black', 'white', 'silver', 'gray', 'grey', 'red', 'blue', 'green', 'brown', 'yellow']
    for color in colors:
        if color in text_lower:
            return color
    return 'unknown'

def extract_weight_capacity(text):
    patterns = [
        r'(\d+)\s*(?:lbs?|pounds?)',
        r'(\d+)\s*(?:kg)',
        r'capacity[:\s]+(\d+)',
        r'holds?\s+(\d+)',
        r'supports?\s+(\d+)'
    ]
    for pattern in patterns:
        match = re.search(pattern, text.lower())
        if match:
            weight = int(match.group(1))
            if 'kg' in pattern:
                weight = int(weight * 2.20462)
            return weight
    return None

def is_rail_or_slatwall(text):
    text_lower = text.lower()
    keywords = ['slatwall', 'slat wall', 'rail system', 'track system', 'french cleat', 'wall track', 'geartrack', 'versatrack', 'fasttrack']
    return 'yes' if any(kw in text_lower for kw in keywords) else 'no'

def get_rail_type(text):
    text_lower = text.lower()
    if is_rail_or_slatwall(text) == 'no':
        return 'none'
    elif 'rail' in text_lower or 'track' in text_lower:
        return 'rail'
    elif 'attachment' in text_lower or 'accessory' in text_lower:
        return 'attachment'
    elif 'kit' in text_lower or 'system' in text_lower:
        return 'kit'
    return 'rail'

def is_hook_or_hanger(text):
    text_lower = text.lower()
    keywords = ['hook', 'hanger', 'hang', 'holder', 'mount']
    return 'yes' if any(kw in text_lower for kw in keywords) else 'no'

# Load ALL products from each retailer
all_products = []
retailers = ['amazon', 'walmart', 'homedepot', 'target', 'lowes']

for retailer in retailers:
    file_path = Path(f'data/retailers/{retailer}_products.json')
    if not file_path.exists():
        print(f"⚠️  {retailer.upper()}: File not found")
        continue

    with open(file_path, 'r', encoding='utf-8') as f:
        products = json.load(f)

    print(f"Loading {len(products)} products from {retailer.upper()}...")

    # Ensure retailer field is set correctly
    for product in products:
        if not product.get('retailer') or product['retailer'].lower() == 'unknown':
            product['retailer'] = retailer.title()
        else:
            product['retailer'] = product['retailer'].title()

        # Apply enhanced fields if not present
        if 'material' not in product or not product.get('material'):
            name = product.get('name', product.get('title', ''))
            description = str(product.get('attributes', ''))
            full_text = f"{name} {description}"

            product['material'] = extract_material(full_text)
            product['color'] = extract_color(full_text)
            product['weight_capacity_lbs'] = extract_weight_capacity(full_text)
            product['is_rail_or_slatwall'] = is_rail_or_slatwall(full_text)
            product['rail_type'] = get_rail_type(full_text)
            product['is_hook_or_hanger'] = is_hook_or_hanger(full_text)

    all_products.extend(products)

print()
print(f"✅ Total products loaded: {len(all_products)}")
print()

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "Category Data"

# Headers
headers = [
    "Retailer",
    "Product Name",
    "Brand",
    "Product Link",
    "Image URL",
    "Price",
    "Star Rating",
    "Review Count",
    "Sales Rank/BSR",
    "Category",
    "Subcategory",
    "Description",
    "Material",
    "Color",
    "Weight Capacity (lbs)",
    "Rail/Slatwall System",
    "Rail Type",
    "Hook/Hanger Product"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write data
for row_num, product in enumerate(all_products, 2):
    ws.cell(row=row_num, column=1).value = product.get('retailer', '').title()
    ws.cell(row=row_num, column=2).value = product.get('name', product.get('title', ''))
    ws.cell(row=row_num, column=3).value = product.get('brand', product.get('shop', ''))

    # Product Link
    url = product.get('url', '')
    ws.cell(row=row_num, column=4).value = url
    if url:
        ws.cell(row=row_num, column=4).style = "Hyperlink"

    # Image URL
    image_url = product.get('image', '')
    ws.cell(row=row_num, column=5).value = image_url
    if image_url:
        ws.cell(row=row_num, column=5).style = "Hyperlink"

    # Price
    price = product.get('price', 0)
    if isinstance(price, (int, float)):
        ws.cell(row=row_num, column=6).value = price
        ws.cell(row=row_num, column=6).number_format = '"$"#,##0.00'
    else:
        try:
            ws.cell(row=row_num, column=6).value = float(str(price).replace('$', '').replace(',', ''))
            ws.cell(row=row_num, column=6).number_format = '"$"#,##0.00'
        except:
            ws.cell(row=row_num, column=6).value = str(price)

    # Rating
    rating = product.get('rating', 0)
    if rating:
        try:
            ws.cell(row=row_num, column=7).value = float(rating)
            ws.cell(row=row_num, column=7).number_format = '0.0'
        except:
            ws.cell(row=row_num, column=7).value = rating

    # Reviews
    reviews = product.get('reviews', product.get('total_reviews', 0))
    if reviews:
        try:
            ws.cell(row=row_num, column=8).value = int(reviews)
            ws.cell(row=row_num, column=8).number_format = '#,##0'
        except:
            ws.cell(row=row_num, column=8).value = reviews

    # BSR
    ws.cell(row=row_num, column=9).value = product.get('bsr', '')

    # Category/Subcategory
    taxonomy = product.get('taxonomy_path', [])
    if isinstance(taxonomy, str):
        try:
            taxonomy = eval(taxonomy)
        except:
            taxonomy = []

    if taxonomy:
        ws.cell(row=row_num, column=10).value = taxonomy[0] if len(taxonomy) > 0 else ''
        ws.cell(row=row_num, column=11).value = taxonomy[1] if len(taxonomy) > 1 else ''

    # Description
    ws.cell(row=row_num, column=12).value = product.get('description', '')

    # ENHANCED FIELDS
    ws.cell(row=row_num, column=13).value = product.get('material', '')
    ws.cell(row=row_num, column=14).value = product.get('color', '')
    ws.cell(row=row_num, column=15).value = product.get('weight_capacity_lbs', '')
    ws.cell(row=row_num, column=16).value = product.get('is_rail_or_slatwall', '')
    ws.cell(row=row_num, column=17).value = product.get('rail_type', '')
    ws.cell(row=row_num, column=18).value = product.get('is_hook_or_hanger', '')

# Column widths
column_widths = {
    1: 12, 2: 50, 3: 20, 4: 15, 5: 15, 6: 12, 7: 12, 8: 12, 9: 15,
    10: 20, 11: 20, 12: 40, 13: 15, 14: 12, 15: 15, 16: 18, 17: 15, 18: 18
}

for col_num, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Freeze header
ws.freeze_panes = "A2"

# Save
output_file = Path("04_CATEGORY_DATA_ALL_PRODUCTS.xlsx")
wb.save(output_file)

print(f"✅ Excel created: {output_file}")
print(f"✅ Total rows: {len(all_products)}")
print()

# Breakdown by retailer
by_retailer = {}
for p in all_products:
    retailer = p.get('retailer', 'Unknown')
    by_retailer[retailer] = by_retailer.get(retailer, 0) + 1

print("="*70)
print("PRODUCT BREAKDOWN BY RETAILER")
print("="*70)
for retailer, count in sorted(by_retailer.items()):
    print(f"  {retailer}: {count:,} products")

print()
print("="*70)
print("COMPLETE DELIVERABLE WITH ALL SCRAPED PRODUCTS")
print("="*70)
