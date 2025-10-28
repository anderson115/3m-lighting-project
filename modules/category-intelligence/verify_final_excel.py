#!/usr/bin/env python3
"""Verify the final Excel file has proper data."""
from openpyxl import load_workbook

wb = load_workbook('04_CATEGORY_DATA_FINAL.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")
print()

# Check for blank product names
blank_names = 0
blank_brands = 0
blank_prices = 0

for row in range(2, ws.max_row + 1):
    if not ws.cell(row, 2).value:  # Product Name
        blank_names += 1
    if not ws.cell(row, 3).value:  # Brand
        blank_brands += 1
    if not ws.cell(row, 6).value:  # Price
        blank_prices += 1

print(f"Blank Product Names: {blank_names}/{ws.max_row-1}")
print(f"Blank Brands: {blank_brands}/{ws.max_row-1}")
print(f"Blank Prices: {blank_prices}/{ws.max_row-1}")
print()

# Show first 3 products
print("Sample products:")
for row in range(2, min(5, ws.max_row + 1)):
    retailer = ws.cell(row, 1).value
    product = ws.cell(row, 2).value
    brand = ws.cell(row, 3).value
    price = ws.cell(row, 6).value
    rating = ws.cell(row, 7).value
    reviews = ws.cell(row, 8).value

    print(f"  Row {row}:")
    print(f"    Retailer: {retailer}")
    print(f"    Product: {product[:60] if product else 'N/A'}...")
    print(f"    Brand: {brand}")
    print(f"    Price: ${price}")
    print(f"    Rating: {rating} ({reviews} reviews)")
    print()
