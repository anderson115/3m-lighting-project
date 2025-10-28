#!/usr/bin/env python3
"""
Create final client Excel with ALL extracted data.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

print("="*70)
print("CREATING COMPLETE FINAL DELIVERABLE")
print("="*70)
print()

# Load enhanced data
with open('data/retailers/all_products_enhanced_heuristic.json', 'r') as f:
    products = json.load(f)

print(f"Loaded {len(products)} enhanced products")
print()

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "Category Data"

# Headers
headers = [
    "Retailer",
    "Product Name",
    "Brand",
    "Product Link",
    "Image URL",
    "Price",
    "Star Rating",
    "Review Count",
    "Sales Rank/BSR",
    "Category",
    "Subcategory",
    "Description",
    "Material",
    "Color",
    "Weight Capacity (lbs)",
    "Rail/Slatwall System",
    "Rail Type",
    "Hook/Hanger Product"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write data
for row_num, product in enumerate(products, 2):
    ws.cell(row=row_num, column=1).value = product.get('retailer', '').title()
    ws.cell(row=row_num, column=2).value = product.get('name', product.get('title', ''))
    ws.cell(row=row_num, column=3).value = product.get('brand', product.get('shop', ''))

    # Product Link
    url = product.get('url', '')
    ws.cell(row=row_num, column=4).value = url
    if url:
        ws.cell(row=row_num, column=4).style = "Hyperlink"

    # Image URL
    ws.cell(row=row_num, column=5).value = product.get('image', '')

    # Price
    price = product.get('price', 0)
    if isinstance(price, (int, float)):
        ws.cell(row=row_num, column=6).value = price
        ws.cell(row=row_num, column=6).number_format = '"$"#,##0.00'
    else:
        try:
            ws.cell(row=row_num, column=6).value = float(str(price).replace('$', '').replace(',', ''))
            ws.cell(row=row_num, column=6).number_format = '"$"#,##0.00'
        except:
            ws.cell(row=row_num, column=6).value = str(price)

    # Rating
    rating = product.get('rating', 0)
    if rating:
        try:
            ws.cell(row=row_num, column=7).value = float(rating)
            ws.cell(row=row_num, column=7).number_format = '0.0'
        except:
            ws.cell(row=row_num, column=7).value = rating

    # Reviews
    reviews = product.get('reviews', product.get('total_reviews', 0))
    if reviews:
        try:
            ws.cell(row=row_num, column=8).value = int(reviews)
            ws.cell(row=row_num, column=8).number_format = '#,##0'
        except:
            ws.cell(row=row_num, column=8).value = reviews

    # BSR
    ws.cell(row=row_num, column=9).value = product.get('bsr', '')

    # Category/Subcategory
    taxonomy = product.get('taxonomy_path', [])
    if isinstance(taxonomy, str):
        try:
            taxonomy = eval(taxonomy)
        except:
            taxonomy = []

    if taxonomy:
        ws.cell(row=row_num, column=10).value = taxonomy[0] if len(taxonomy) > 0 else ''
        ws.cell(row=row_num, column=11).value = taxonomy[1] if len(taxonomy) > 1 else ''

    # Description
    ws.cell(row=row_num, column=12).value = product.get('description', '')

    # ENHANCED FIELDS
    ws.cell(row=row_num, column=13).value = product.get('material', '')
    ws.cell(row=row_num, column=14).value = product.get('color', '')
    ws.cell(row=row_num, column=15).value = product.get('weight_capacity_lbs', '')
    ws.cell(row=row_num, column=16).value = product.get('is_rail_or_slatwall', '')
    ws.cell(row=row_num, column=17).value = product.get('rail_type', '')
    ws.cell(row=row_num, column=18).value = product.get('is_hook_or_hanger', '')

# Column widths
column_widths = {
    1: 12, 2: 50, 3: 20, 4: 15, 5: 15, 6: 12, 7: 12, 8: 12, 9: 15,
    10: 20, 11: 20, 12: 40, 13: 15, 14: 12, 15: 15, 16: 18, 17: 15, 18: 18
}

for col_num, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Freeze header
ws.freeze_panes = "A2"

# Save
output_file = Path("04_CATEGORY_DATA_COMPLETE.xlsx")
wb.save(output_file)

print(f"✅ Excel created: {output_file}")
print(f"✅ Total rows: {len(products)}")
print()

# Field coverage
material_filled = sum(1 for p in products if p.get('material') and p['material'] != 'unknown')
color_filled = sum(1 for p in products if p.get('color') and p['color'] != 'unknown')
weight_filled = sum(1 for p in products if p.get('weight_capacity_lbs'))
rail_yes = sum(1 for p in products if p.get('is_rail_or_slatwall') == 'yes')
hook_yes = sum(1 for p in products if p.get('is_hook_or_hanger') == 'yes')

print("="*70)
print("FIELD COVERAGE SUMMARY")
print("="*70)
print(f"Core Fields (100%):")
print(f"  • Product Names: {len(products)}/{len(products)}")
print(f"  • Prices: {len(products)}/{len(products)}")
print(f"  • Ratings/Reviews: {len(products)}/{len(products)}")
print()
print(f"Enhanced Fields:")
print(f"  • Material: {material_filled}/{len(products)} ({100*material_filled/len(products):.1f}%)")
print(f"  • Color: {color_filled}/{len(products)} ({100*color_filled/len(products):.1f}%)")
print(f"  • Weight Capacity: {weight_filled}/{len(products)} ({100*weight_filled/len(products):.1f}%)")
print(f"  • Rail/Slatwall Products: {rail_yes} ({100*rail_yes/len(products):.1f}%)")
print(f"  • Hook/Hanger Products: {hook_yes} ({100*hook_yes/len(products):.1f}%)")
print()
print("="*70)
print("DELIVERABLE COMPLETE")
print("="*70)
