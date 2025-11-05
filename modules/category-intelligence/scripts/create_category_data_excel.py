#!/usr/bin/env python3
"""
Create 04_CATEGORY_DATA.XLSX with all scraped retailer product data.
Formats columns properly and includes all requested fields.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Define data directory
DATA_DIR = Path(__file__).parent / "data" / "retailers"

print("=" * 70)
print("CREATING 04_CATEGORY_DATA.XLSX")
print("=" * 70)
print()

# Load all retailer data
all_products = []

# Amazon
print("Loading Amazon products...")
try:
    with open(DATA_DIR / "amazon_products.json", 'r', encoding='utf-8') as f:
        amazon_data = json.load(f)
        for product in amazon_data:
            all_products.append({
                'Retailer': 'Amazon',
                'Product ID': product.get('asin', ''),
                'Product Name': product.get('title', ''),
                'Brand': product.get('brand', ''),
                'Category': product.get('category', ''),
                'Subcategory': product.get('subcategory', ''),
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('reviews', product.get('total_reviews', 0)),
                'Description': product.get('description', product.get('features', '')),
                'Product URL': f"https://www.amazon.com/dp/{product.get('asin', '')}" if product.get('asin') else '',
                'Image URL': product.get('image', product.get('image_url', '')),
                'BSR': product.get('bsr'),
                'Est. Monthly Sales': product.get('estimated_monthly_sales')
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Amazon'])} Amazon products")
except Exception as e:
    print(f"  ✗ Error loading Amazon: {e}")

# Walmart
print("Loading Walmart products...")
try:
    with open(DATA_DIR / "walmart_products.json", 'r', encoding='utf-8') as f:
        walmart_data = json.load(f)
        for product in walmart_data:
            all_products.append({
                'Retailer': 'Walmart',
                'Product ID': product.get('product_id', product.get('id', '')),
                'Product Name': product.get('name', product.get('title', '')),
                'Brand': product.get('brand', ''),
                'Category': product.get('category', ''),
                'Subcategory': product.get('subcategory', ''),
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('review_count', product.get('reviews', 0)),
                'Description': product.get('description', ''),
                'Product URL': product.get('url', product.get('product_url', '')),
                'Image URL': product.get('image', product.get('image_url', '')),
                'BSR': None,
                'Est. Monthly Sales': None
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Walmart'])} Walmart products")
except Exception as e:
    print(f"  ✗ Error loading Walmart: {e}")

# Home Depot
print("Loading Home Depot products...")
try:
    with open(DATA_DIR / "homedepot_products.json", 'r', encoding='utf-8') as f:
        homedepot_data = json.load(f)
        for product in homedepot_data:
            all_products.append({
                'Retailer': 'Home Depot',
                'Product ID': product.get('product_id', product.get('sku', '')),
                'Product Name': product.get('name', product.get('title', '')),
                'Brand': product.get('brand', ''),
                'Category': product.get('category', ''),
                'Subcategory': product.get('subcategory', ''),
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('review_count', product.get('reviews', 0)),
                'Description': product.get('description', ''),
                'Product URL': product.get('url', product.get('product_url', '')),
                'Image URL': product.get('image', product.get('image_url', '')),
                'BSR': None,
                'Est. Monthly Sales': None
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Home Depot'])} Home Depot products")
except Exception as e:
    print(f"  ✗ Error loading Home Depot: {e}")

# Target
print("Loading Target products...")
try:
    with open(DATA_DIR / "target_products.json", 'r', encoding='utf-8') as f:
        target_data = json.load(f)
        for product in target_data:
            all_products.append({
                'Retailer': 'Target',
                'Product ID': product.get('product_id', product.get('tcin', '')),
                'Product Name': product.get('name', product.get('title', '')),
                'Brand': product.get('brand', ''),
                'Category': product.get('category', ''),
                'Subcategory': product.get('subcategory', ''),
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('review_count', product.get('reviews', 0)),
                'Description': product.get('description', ''),
                'Product URL': product.get('url', product.get('product_url', '')),
                'Image URL': product.get('image', product.get('image_url', '')),
                'BSR': None,
                'Est. Monthly Sales': None
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Target'])} Target products")
except Exception as e:
    print(f"  ✗ Error loading Target: {e}")

# Etsy
print("Loading Etsy products...")
try:
    with open(DATA_DIR / "etsy_products.json", 'r', encoding='utf-8') as f:
        etsy_data = json.load(f)
        for product in etsy_data:
            all_products.append({
                'Retailer': 'Etsy',
                'Product ID': str(product.get('listing_id', product.get('id', ''))),
                'Product Name': product.get('title', product.get('name', '')),
                'Brand': product.get('shop_name', product.get('brand', '')),
                'Category': product.get('category', ''),
                'Subcategory': product.get('subcategory', ''),
                'Price': product.get('price'),
                'Rating': product.get('rating'),
                'Review Count': product.get('review_count', product.get('reviews', 0)),
                'Description': product.get('description', ''),
                'Product URL': product.get('url', product.get('product_url', '')),
                'Image URL': product.get('image', product.get('image_url', '')),
                'BSR': None,
                'Est. Monthly Sales': None
            })
    print(f"  ✓ Loaded {len([p for p in all_products if p['Retailer'] == 'Etsy'])} Etsy products")
except Exception as e:
    print(f"  ✗ Error loading Etsy: {e}")

print()
print(f"Total products loaded: {len(all_products)}")
print()

# Create DataFrame
df = pd.DataFrame(all_products)

# Reorder columns
column_order = [
    'Retailer',
    'Product ID',
    'Product Name',
    'Brand',
    'Category',
    'Subcategory',
    'Price',
    'Rating',
    'Review Count',
    'Description',
    'Product URL',
    'Image URL',
    'BSR',
    'Est. Monthly Sales'
]

df = df[column_order]

# Format data types
df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
df['Rating'] = pd.to_numeric(df['Rating'], errors='coerce')
df['Review Count'] = pd.to_numeric(df['Review Count'], errors='coerce').fillna(0).astype(int)
df['BSR'] = pd.to_numeric(df['BSR'], errors='coerce')
df['Est. Monthly Sales'] = pd.to_numeric(df['Est. Monthly Sales'], errors='coerce')

# Sort by Retailer and then by Product Name
df = df.sort_values(['Retailer', 'Product Name'])

# Save to Excel
output_file = Path(__file__).parent / "04_CATEGORY_DATA.xlsx"
print(f"Saving to Excel: {output_file.name}")
df.to_excel(output_file, index=False, sheet_name='Product Data')

# Format Excel file
print("Applying Excel formatting...")
wb = load_workbook(output_file)
ws = wb['Product Data']

# Header formatting
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
column_widths = {
    'A': 12,  # Retailer
    'B': 15,  # Product ID
    'C': 50,  # Product Name
    'D': 20,  # Brand
    'E': 25,  # Category
    'F': 25,  # Subcategory
    'G': 10,  # Price
    'H': 8,   # Rating
    'I': 12,  # Review Count
    'J': 60,  # Description
    'K': 50,  # Product URL
    'L': 50,  # Image URL
    'M': 8,   # BSR
    'N': 15   # Est. Monthly Sales
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Format price columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row:
        if cell.value:
            cell.number_format = '$#,##0.00'

# Format rating columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
    for cell in row:
        if cell.value:
            cell.number_format = '0.0'

# Format review count columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
    for cell in row:
        if cell.value:
            cell.number_format = '#,##0'

# Format BSR columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=13, max_col=13):
    for cell in row:
        if cell.value:
            cell.number_format = '#,##0'

# Format Est. Monthly Sales columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=14, max_col=14):
    for cell in row:
        if cell.value:
            cell.number_format = '#,##0'

# Freeze header row
ws.freeze_panes = 'A2'

# Add filters
ws.auto_filter.ref = ws.dimensions

# Save formatted workbook
wb.save(output_file)

print()
print("=" * 70)
print("EXCEL FILE CREATED SUCCESSFULLY")
print("=" * 70)
print(f"File: {output_file.name}")
print(f"Total products: {len(df):,}")
print()
print("Breakdown by retailer:")
for retailer in df['Retailer'].unique():
    count = len(df[df['Retailer'] == retailer])
    print(f"  {retailer}: {count:,} products")
print()
print("Columns included:")
for col in column_order:
    print(f"  - {col}")
print()
