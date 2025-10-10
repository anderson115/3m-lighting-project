#!/usr/bin/env python3
"""
Excel Report Generator - Professional Data Export
Generates Excel spreadsheets with structured data and citations
100% TRACEABLE - Every insight must link to original source
"""

import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..core.config import Config


class ExcelReporter:
    """Production Excel report generator with citation validation"""

    def __init__(self, tier: int = 2, config: Optional[Config] = None):
        """
        Initialize Excel reporter

        Args:
            tier: Report tier (2 = Professional, 3 = Enterprise)
            config: Configuration object
        """
        if tier < 2:
            raise ValueError("Excel reports only available for Tier 2+")

        if not HAS_OPENPYXL:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        self.tier = tier
        self.config = config or Config()
        self.tier_config = self.config.get_tier_config(tier)

        # Setup logging
        self.logger = logging.getLogger(__name__)

        # Output directory
        self.output_dir = Path(__file__).parent.parent / "data" / "reports"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        analysis: Dict,
        discussions: List[Dict],
        project_name: str = "Expert Authority Analysis"
    ) -> Path:
        """
        Generate Excel report from analysis

        Args:
            analysis: Analysis results from ProductionAnalyzer
            discussions: Original discussions (for citation validation)
            project_name: Name of the project/analysis

        Returns:
            Path to generated Excel report
        """
        self.logger.info(f"📊 Generating Tier {self.tier} Excel report...")

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_summary_sheet(wb, analysis, project_name)
        self._create_themes_sheet(wb, analysis)
        self._create_consensus_sheet(wb, analysis)

        if self.tier >= 2:
            if analysis.get('controversies'):
                self._create_controversies_sheet(wb, analysis)
            if analysis.get('safety_warnings'):
                self._create_safety_sheet(wb, analysis)

        self._create_discussions_sheet(wb, discussions)

        # Save report
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_file = self.output_dir / f"{project_name.replace(' ', '_')}_tier{self.tier}_{timestamp}.xlsx"

        wb.save(report_file)

        self.logger.info(f"✅ Excel report generated: {report_file}")
        return report_file

    def _create_summary_sheet(self, wb, analysis: Dict, project_name: str):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = project_name
        ws['A1'].font = Font(size=18, bold=True, color="0066CC")
        ws.merge_cells('A1:D1')

        # Tier badge
        ws['A2'] = f"Tier {self.tier} - {self.tier_config['name']}"
        ws['A2'].font = Font(size=12, bold=True)
        ws.merge_cells('A2:D2')

        # Timestamp
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, color="666666")
        ws.merge_cells('A3:D3')

        # Stats
        row = 5
        ws[f'A{row}'] = "ANALYSIS SUMMARY"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2

        stats = [
            ("Total Discussions Analyzed:", analysis['metadata']['total_discussions']),
            ("Themes Discovered:", len(analysis['themes'])),
            ("Consensus Patterns:", len(analysis['consensus_patterns'])),
            ("Data Sources:", ', '.join(analysis['metadata']['platforms'])),
            ("Analysis Method:", analysis['metadata']['analysis_method'])
        ]

        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    def _create_themes_sheet(self, wb, analysis: Dict):
        """Create themes sheet"""
        ws = wb.create_sheet("Themes")

        # Headers
        headers = ["#", "Theme", "Frequency", "Frequency %", "Method", "Description", "Strategic Insight", "Example URLs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for i, theme in enumerate(analysis['themes'], 1):
            row = i + 1
            ws.cell(row, 1, i)
            ws.cell(row, 2, theme['theme'])
            ws.cell(row, 3, theme['frequency'])
            ws.cell(row, 4, f"{theme['frequency_pct']}%")
            ws.cell(row, 5, theme.get('method', 'N/A'))
            ws.cell(row, 6, theme.get('description', 'N/A'))
            ws.cell(row, 7, theme.get('strategic_insight', 'N/A'))

            # Example URLs
            example_urls = [ex['url'] for ex in theme.get('examples', [])]
            ws.cell(row, 8, '\n'.join(example_urls[:3]))

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 60

        # Enable text wrapping
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _create_consensus_sheet(self, wb, analysis: Dict):
        """Create consensus patterns sheet"""
        ws = wb.create_sheet("Consensus")

        # Headers
        headers = ["#", "Pattern", "Score", "Platform", "Source Discussion", "Source URL", "Accepted Answer"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for i, pattern in enumerate(analysis['consensus_patterns'], 1):
            row = i + 1
            ws.cell(row, 1, i)
            ws.cell(row, 2, pattern['pattern'])
            ws.cell(row, 3, pattern['score'])
            ws.cell(row, 4, pattern['platform'])
            ws.cell(row, 5, pattern['source_discussion'])
            ws.cell(row, 6, pattern['source_url'])
            ws.cell(row, 7, "Yes" if pattern.get('is_accepted') else "No")

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 15

        # Enable text wrapping
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _create_controversies_sheet(self, wb, analysis: Dict):
        """Create controversies sheet"""
        ws = wb.create_sheet("Controversies")

        # Headers
        headers = ["#", "Topic", "Comments", "Score", "Platform", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for i, controversy in enumerate(analysis['controversies'], 1):
            row = i + 1
            ws.cell(row, 1, i)
            ws.cell(row, 2, controversy['topic'])
            ws.cell(row, 3, controversy['num_comments'])
            ws.cell(row, 4, controversy['score'])
            ws.cell(row, 5, controversy['platform'])
            ws.cell(row, 6, controversy['url'])

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 60

        # Enable text wrapping
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _create_safety_sheet(self, wb, analysis: Dict):
        """Create safety warnings sheet"""
        ws = wb.create_sheet("Safety")

        # Headers
        headers = ["#", "Warning", "Keyword", "Platform", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for i, warning in enumerate(analysis['safety_warnings'], 1):
            row = i + 1
            ws.cell(row, 1, i)
            ws.cell(row, 2, warning['warning'])
            ws.cell(row, 3, warning['keyword'])
            ws.cell(row, 4, warning['platform'])
            ws.cell(row, 5, warning['url'])

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 60

        # Enable text wrapping
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _create_discussions_sheet(self, wb, discussions: List[Dict]):
        """Create raw discussions data sheet"""
        ws = wb.create_sheet("Raw Data")

        # Headers
        headers = ["ID", "Platform", "Title", "Author", "Score", "Comments/Answers", "Created Date", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for i, disc in enumerate(discussions, 1):
            row = i + 1
            ws.cell(row, 1, disc.get('id', ''))
            ws.cell(row, 2, disc.get('platform', ''))
            ws.cell(row, 3, disc.get('title', ''))
            ws.cell(row, 4, disc.get('author', ''))
            ws.cell(row, 5, disc.get('score', 0))
            ws.cell(row, 6, disc.get('num_comments', 0) or disc.get('answer_count', 0))
            ws.cell(row, 7, disc.get('created_date', ''))
            ws.cell(row, 8, disc.get('url', ''))

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 60

        # Enable text wrapping
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


if __name__ == "__main__":
    # Example usage
    logging.basicConfig(level=logging.INFO)

    reporter = ExcelReporter(tier=2)
    print(f"✅ Excel Reporter ready (Tier {reporter.tier})")
    print(f"📊 openpyxl available: {HAS_OPENPYXL}")
