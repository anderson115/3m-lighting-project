#!/usr/bin/env python3
"""
Create master product database spreadsheet with:
- Executive Summary tab (retailer, category, brand coverage)
- Master tab with all products combined
- Separate tabs for each retailer
"""
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("="*60)
print("CREATING MASTER PRODUCT DATABASE SPREADSHEET")
print("="*60)

# Load product data
print("\n1. Loading product data...")
with open('06-final-deliverables/13-raw-data/all_products_final_with_lowes.json') as f:
    all_products = json.load(f)

print(f"   ✓ Loaded {len(all_products)} total products")

# Convert to DataFrame
df = pd.DataFrame(all_products)
print(f"   ✓ Converted to DataFrame: {df.shape[0]} rows × {df.shape[1]} columns")

# Standardize column names
column_mapping = {
    'product_name': 'name',
    'product_url': 'url',
    'review_count': 'reviews'
}
for old, new in column_mapping.items():
    if old in df.columns and new not in df.columns:
        df[new] = df[old]

# Create workbook
print("\n2. Creating Excel workbook...")
wb = Workbook()
wb.remove(wb.active)

# ============================================
# EXECUTIVE SUMMARY TAB
# ============================================
print("\n3. Creating Executive Summary tab...")
exec_sheet = wb.create_sheet("Executive Summary", 0)

# Title
exec_sheet.merge_cells('A1:F1')
exec_sheet['A1'] = "3M GARAGE ORGANIZATION - MASTER PRODUCT DATABASE"
exec_sheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
exec_sheet['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
exec_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
exec_sheet.row_dimensions[1].height = 30

# Subtitle
exec_sheet.merge_cells('A2:F2')
exec_sheet['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
exec_sheet['A2'].font = Font(size=10, italic=True)
exec_sheet['A2'].alignment = Alignment(horizontal='center')

# Overview section
exec_sheet['A4'] = "DATABASE OVERVIEW"
exec_sheet['A4'].font = Font(size=14, bold=True, color="1F4E78")
exec_sheet.merge_cells('A4:B4')

exec_sheet['A5'] = "Total Products:"
exec_sheet['B5'] = len(df)
exec_sheet['A6'] = "Unique Retailers:"
exec_sheet['B6'] = df['retailer'].nunique() if 'retailer' in df.columns else 0
exec_sheet['A7'] = "Unique Brands:"
exec_sheet['B7'] = df['brand'].nunique() if 'brand' in df.columns else 0
exec_sheet['A8'] = "Unique Categories:"
exec_sheet['B8'] = df['category'].nunique() if 'category' in df.columns else 0
exec_sheet['A9'] = "Data Collection Period:"
exec_sheet['B9'] = "March - November 2025"

for row in range(5, 10):
    exec_sheet[f'A{row}'].font = Font(bold=True)

# Retailer breakdown
exec_sheet['A11'] = "RETAILER COVERAGE"
exec_sheet['A11'].font = Font(size=12, bold=True, color="1F4E78")
exec_sheet.merge_cells('A11:D11')

exec_sheet['A12'] = "Retailer"
exec_sheet['B12'] = "Product Count"
exec_sheet['C12'] = "% of Total"
exec_sheet['D12'] = "Avg Price"

for col in ['A12', 'B12', 'C12', 'D12']:
    exec_sheet[col].font = Font(bold=True, color="FFFFFF")
    exec_sheet[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    exec_sheet[col].alignment = Alignment(horizontal='center')

if 'retailer' in df.columns:
    retailer_stats = df.groupby('retailer').agg({
        'name': 'count',
        'price': 'mean'
    }).reset_index()
    retailer_stats.columns = ['Retailer', 'Count', 'Avg Price']
    retailer_stats['Percentage'] = (retailer_stats['Count'] / len(df) * 100).round(1)
    retailer_stats = retailer_stats.sort_values('Count', ascending=False)

    row = 13
    for _, r in retailer_stats.iterrows():
        exec_sheet[f'A{row}'] = r['Retailer']
        exec_sheet[f'B{row}'] = r['Count']
        exec_sheet[f'C{row}'] = f"{r['Percentage']}%"
        exec_sheet[f'D{row}'] = f"${r['Avg Price']:.2f}" if pd.notna(r['Avg Price']) else 'N/A'
        row += 1

    print(f"   ✓ Added {len(retailer_stats)} retailers")

# Category breakdown
cat_start = row + 2
exec_sheet[f'A{cat_start}'] = "TOP 15 CATEGORIES"
exec_sheet[f'A{cat_start}'].font = Font(size=12, bold=True, color="1F4E78")
exec_sheet.merge_cells(f'A{cat_start}:C{cat_start}')

exec_sheet[f'A{cat_start+1}'] = "Category"
exec_sheet[f'B{cat_start+1}'] = "Product Count"
exec_sheet[f'C{cat_start+1}'] = "% of Total"

for col in [f'A{cat_start+1}', f'B{cat_start+1}', f'C{cat_start+1}']:
    exec_sheet[col].font = Font(bold=True, color="FFFFFF")
    exec_sheet[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

if 'category' in df.columns:
    categories = df['category'].value_counts().head(15)
    row = cat_start + 2
    for cat, count in categories.items():
        exec_sheet[f'A{row}'] = cat
        exec_sheet[f'B{row}'] = count
        exec_sheet[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1

    print(f"   ✓ Added top 15 categories")

# Brand breakdown
brand_start = cat_start + 18
exec_sheet[f'A{brand_start}'] = "TOP 15 BRANDS"
exec_sheet[f'A{brand_start}'].font = Font(size=12, bold=True, color="1F4E78")
exec_sheet.merge_cells(f'A{brand_start}:C{brand_start}')

exec_sheet[f'A{brand_start+1}'] = "Brand"
exec_sheet[f'B{brand_start+1}'] = "Product Count"
exec_sheet[f'C{brand_start+1}'] = "% of Total"

for col in [f'A{brand_start+1}', f'B{brand_start+1}', f'C{brand_start+1}']:
    exec_sheet[col].font = Font(bold=True, color="FFFFFF")
    exec_sheet[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

if 'brand' in df.columns:
    brands = df['brand'].value_counts().head(15)
    row = brand_start + 2
    for brand, count in brands.items():
        exec_sheet[f'A{row}'] = brand
        exec_sheet[f'B{row}'] = count
        exec_sheet[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1

    print(f"   ✓ Added top 15 brands")

# Set column widths
exec_sheet.column_dimensions['A'].width = 30
exec_sheet.column_dimensions['B'].width = 15
exec_sheet.column_dimensions['C'].width = 15
exec_sheet.column_dimensions['D'].width = 15

# ============================================
# ALL PRODUCTS TAB
# ============================================
print("\n4. Creating All Products tab...")
all_sheet = wb.create_sheet("All Products Combined")

# Headers
headers = ["Product Name", "Brand", "Price", "Category", "Retailer", "Rating", "Reviews", "URL"]
all_sheet.append(headers)

# Style headers
for col_num, header in enumerate(headers, 1):
    cell = all_sheet.cell(row=1, column=col_num)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for idx, row in df.iterrows():
    all_sheet.append([
        row.get('name', ''),
        row.get('brand', ''),
        row.get('price', ''),
        row.get('category', ''),
        row.get('retailer', ''),
        row.get('rating', ''),
        row.get('reviews', ''),
        row.get('url', '')
    ])

    if (idx + 1) % 2000 == 0:
        print(f"   → Processed {idx + 1:,} products...")

# Set column widths
all_sheet.column_dimensions['A'].width = 50
all_sheet.column_dimensions['B'].width = 20
all_sheet.column_dimensions['C'].width = 12
all_sheet.column_dimensions['D'].width = 25
all_sheet.column_dimensions['E'].width = 15
all_sheet.column_dimensions['F'].width = 10
all_sheet.column_dimensions['G'].width = 10
all_sheet.column_dimensions['H'].width = 60

print(f"   ✓ Added {len(df):,} products to master tab")

# ============================================
# RETAILER-SPECIFIC TABS
# ============================================
print("\n5. Creating retailer-specific tabs...")

if 'retailer' in df.columns:
    # Filter out NaN values and convert all to strings to avoid sorting errors
    retailers = df['retailer'].dropna().unique()
    retailers = [str(r) for r in retailers if str(r) != 'nan']  # Extra safety check

    for retailer in sorted(retailers):
        # Clean retailer name for sheet
        sheet_name = str(retailer).replace('/', '-')[:31]

        retailer_df = df[df['retailer'] == retailer].copy()

        if len(retailer_df) == 0:
            continue

        # Create sheet
        sheet = wb.create_sheet(sheet_name)

        # Add headers
        sheet.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data
        for idx, row in retailer_df.iterrows():
            sheet.append([
                row.get('name', ''),
                row.get('brand', ''),
                row.get('price', ''),
                row.get('category', ''),
                row.get('retailer', ''),
                row.get('rating', ''),
                row.get('reviews', ''),
                row.get('url', '')
            ])

        # Set column widths
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 60

        print(f"   ✓ Created '{sheet_name}' tab ({len(retailer_df):,} products)")

# Save workbook
output_path = '06-final-deliverables/08-MASTER_PRODUCT_DATABASE.xlsx'
print(f"\n6. Saving workbook...")
wb.save(output_path)

print("\n" + "="*60)
print("✅ MASTER PRODUCT DATABASE CREATED")
print("="*60)
print(f"File: {output_path}")
print(f"Total Products: {len(df):,}")
print(f"Tabs Created: {len(wb.sheetnames)}")
print(f"Sheets: {', '.join(wb.sheetnames[:5])}{'...' if len(wb.sheetnames) > 5 else ''}")
print("="*60)
