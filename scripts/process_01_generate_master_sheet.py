#!/usr/bin/env python3
"""Build the consolidated IG preference spreadsheet with equity + hypothesis signals."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Iterable

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_POSTS = ROOT / 'outputs' / 'manual_sample_posts.json'
DEFAULT_SNIPPETS = ROOT / 'outputs' / 'manual_snippet_annotations.yaml'
DEFAULT_OUTPUT = ROOT / 'outputs' / 'ig_preference_manual_coding.xlsx'

MEASURE_KEYWORDS = {
    'connection': [
        'friend', 'friends', 'community', 'connect', 'connection', 'dm', 'message',
        'send it', 'share it', 'together', 'audience', 'followers', 'fanbase', 'engage',
        'comment', 'conversation', 'collab', 'co-create', 'ugc',
    ],
    'discovery': [
        'discover', 'discovery', 'explore', 'exploration', 'explore page', 'find new',
        'new ideas', 'new content', 'learn new', 'recommend', 'recommendation', 'search',
        'exposure', 'reach new', 'algorithm shows',
    ],
    'expression': [
        'express', 'expression', 'be myself', 'authentic', 'identity', 'voice', 'story',
        'storytelling', 'share my story', 'showcase', 'portfolio', 'aesthetic', 'style',
    ],
    'entertainment': [
        'entertain', 'entertainment', 'fun', 'funny', 'bored', 'enjoy', 'enjoyable', 'laugh',
        'lol', 'meme', 'drama', 'vibe', 'watch',
    ],
    'creativity': [
        'creative', 'creativity', 'create', 'creating', 'creation', 'edit', 'editing', 'filters',
        'effects', 'art', 'design', 'shoot', 'film', 'script', 'idea',
    ],
}

CREATOR_KEYWORDS = [
    'creator', 'influencer', 'artist', 'agency', 'client work', 'campaign', 'brand partner',
    'freelancer', 'coach', 'consultant', 'business page', 'portfolio', 'paid collab',
]

CONTENT_QUALITY_KEYWORDS = [
    'high-quality', 'quality', 'production', 'aesthetic', 'visuals', 'storytelling', 'story',
    'transitions', 'edit', 'editing', 'polish', 'thumb-stopping', 'cinematic', 'sound design',
    'music', 'audio', 'lighting', 'shoot', 'film', 'script', 'concept', 'idea', 'creative',
    'viral format', 'format', 'template', 'quiz', 'interactive', 'experience', 'immersion',
]

HYPOTHESIS_DRIVER_KEYWORDS = {
    'Algorithm_Trust': ['algorithm', 'reach', 'distribution', 'shadowban', 'explore page', 'feed'],
    'Growth_Reach': ['grow', 'growth', 'views', 'impressions', 'followers', 'scale', 'audience'],
    'Content_Quality': ['quality', 'production', 'visuals', 'storytelling', 'format', 'idea'],
    'Monetization_Path': ['monetize', 'ads', 'budget', 'campaign', 'client', 'sell', 'business', 'revenue'],
    'Community_Culture': ['community', 'friend', 'comments', 'together', 'share', 'vibe', 'culture'],
    'Safety_Governance': ['shadowban', 'ban', 'policy', 'safety', 'graphic', 'throttle'],
    'Workflow_UX': ['workflow', 'process', 'template', 'tool', 'experience', 'interface'],
}

FOLLOWER_PATTERN = re.compile(
    r"(?P<num>\d{1,3}(?:[,.]\d{3})*|\d+(?:\.\d+)?)\s*(?P<suffix>[kKmM])?\s*(?:\+)?\s*(followers?|subs(?:cribers)?|fans|audience)"
)


def load_posts(path: Path) -> tuple[pd.DataFrame, list[dict]]:
    data = json.loads(path.read_text())
    frame = pd.DataFrame(
        {
            'post_id': item['post_id'],
            'platform': item['platform'],
            'source_url': item['url'],
            'raw_text': item['text'],
        }
        for item in data
    )
    return frame, data


def load_snippets(path: Path) -> pd.DataFrame:
    data = yaml.safe_load(path.read_text())['snippets']
    rows = []
    for entry in data:
        rows.append(
            {
                'post_id': entry['post_id'],
                'snippet_id': entry['snippet_id'],
                'agent_interpretation': entry['atomic_snippet'].replace('\n', ' ').strip(),
                'agent_rationale': entry.get('insight_note', ''),
                'driver_topic': entry['driver_topic'],
                'sentiment': entry['sentiment'],
                'app_signal': entry['app_signal'],
                'age_inference': entry['age_inference'],
                'review_action': 'Pending',
                'review_notes': '',
            }
        )
    df = pd.DataFrame(rows)
    return df.sort_values(['post_id', 'snippet_id']).reset_index(drop=True)


def select_example_posts(posts: Iterable[dict], limit: int = 10) -> list[str]:
    seen: list[str] = []
    for post in posts:
        pid = post['post_id']
        if pid not in seen:
            seen.append(pid)
        if len(seen) == limit:
            break
    return seen


def tokenize(text: str) -> set[str]:
    cleaned = ''.join(ch.lower() if ch.isalnum() else ' ' for ch in text)
    return {tok for tok in cleaned.split() if tok}


def re_split_sentences(text: str) -> list[str]:
    delimiters = '.!?\n'
    sentence = []
    sentences = []
    for char in text:
        sentence.append(char)
        if char in delimiters:
            sentences.append(''.join(sentence))
            sentence = []
    if sentence:
        sentences.append(''.join(sentence))
    return sentences


def best_verbatim_excerpt(post_text: str, interpretation: str, max_length: int = 360) -> str:
    sentences = [seg.strip() for seg in re_split_sentences(post_text) if seg.strip()]
    if not sentences:
        return post_text[:max_length]
    target = tokenize(interpretation)
    best_sentence = sentences[0]
    best_score = 0
    for sentence in sentences:
        overlap = tokenize(sentence) & target
        score = len(overlap)
        if score > best_score:
            best_sentence = sentence
            best_score = score
    excerpt = best_sentence if best_score > 0 else sentences[0]
    return excerpt[:max_length]


def score_measure(text: str, keywords: list[str]) -> str:
    lowered = text.lower()
    matches = sum(1 for keyword in keywords if keyword in lowered)
    if matches >= 2:
        return 'Strong'
    if matches == 1:
        return 'Possible'
    return 'None'


def detect_established_creator(text: str) -> str:
    for match in FOLLOWER_PATTERN.finditer(text):
        number = match.group('num').replace(',', '')
        suffix = (match.group('suffix') or '').lower()
        try:
            value = float(number)
        except ValueError:
            continue
        if suffix == 'k':
            value *= 1_000
        elif suffix == 'm':
            value *= 1_000_000
        if value >= 10_000:
            return 'Confirmed'
    lowered = text.lower()
    if any(keyword in lowered for keyword in CREATOR_KEYWORDS):
        return 'Possible'
    return 'Unknown'


def detect_hypothesis_drivers(text: str) -> set[str]:
    lowered = text.lower()
    tags: set[str] = set()
    for driver, keywords in HYPOTHESIS_DRIVER_KEYWORDS.items():
        if any(keyword in lowered for keyword in keywords):
            tags.add(driver)
    return tags


def add_verbatims_and_signals(snippets_df: pd.DataFrame, posts: list[dict]) -> pd.DataFrame:
    post_map = {item['post_id']: item for item in posts}
    verbatims = []
    connection = []
    discovery = []
    expression = []
    entertainment = []
    creativity = []
    content_quality = []
    established = []
    hypothesis_tags = []
    for row in snippets_df.itertuples():
        post = post_map[row.post_id]
        excerpt = best_verbatim_excerpt(post['text'], row.agent_interpretation)
        combined_text = f"{row.agent_interpretation} {excerpt} {row.agent_rationale}"
        verbatims.append(excerpt)
        connection.append(score_measure(combined_text, MEASURE_KEYWORDS['connection']))
        discovery.append(score_measure(combined_text, MEASURE_KEYWORDS['discovery']))
        expression.append(score_measure(combined_text, MEASURE_KEYWORDS['expression']))
        entertainment.append(score_measure(combined_text, MEASURE_KEYWORDS['entertainment']))
        creativity.append(score_measure(combined_text, MEASURE_KEYWORDS['creativity']))
        content_quality.append(score_measure(combined_text, CONTENT_QUALITY_KEYWORDS))
        established.append(detect_established_creator(combined_text))
        tags = detect_hypothesis_drivers(combined_text)
        hypothesis_tags.append('; '.join(sorted(tags)) if tags else 'None')

    updated = snippets_df.copy()
    updated.insert(2, 'verbatim_excerpt', verbatims)
    updated['connection_signal'] = connection
    updated['discovery_signal'] = discovery
    updated['expression_signal'] = expression
    updated['entertainment_signal'] = entertainment
    updated['creativity_signal'] = creativity
    updated['content_quality_signal'] = content_quality
    updated['established_creator_flag'] = established
    updated['hypothesis_driver_tags'] = hypothesis_tags
    updated = add_post_level_labels(updated)
    cols = [
        'post_id', 'snippet_id', 'verbatim_excerpt', 'agent_interpretation', 'agent_rationale',
        'driver_topic', 'sentiment', 'app_signal', 'age_inference', 'established_creator_flag',
        'connection_signal', 'discovery_signal', 'expression_signal', 'entertainment_signal',
        'creativity_signal', 'content_quality_signal', 'hypothesis_driver_tags',
        'post_overall_driver', 'post_overall_sentiment', 'post_established_flag',
        'review_action', 'review_notes',
    ]
    return updated[cols]


def add_post_level_labels(snippets_df: pd.DataFrame) -> pd.DataFrame:
    driver_map: dict[str, str] = {}
    sentiment_map: dict[str, str] = {}
    creator_map: dict[str, str] = {}
    sentiment_priority = ['Negative', 'Mixed', 'Neutral', 'Positive']
    sentiment_rank = {value: idx for idx, value in enumerate(sentiment_priority)}
    creator_rank = {'Unknown': 0, 'Possible': 1, 'Confirmed': 2}

    for post_id, group in snippets_df.groupby('post_id'):
        driver_counts = group['driver_topic'].value_counts()
        driver_map[post_id] = driver_counts.idxmax()

        sentiment_counts = group['sentiment'].value_counts()
        max_count = sentiment_counts.max()
        candidates = [s for s, count in sentiment_counts.items() if count == max_count]
        ordered = sorted(candidates, key=lambda x: sentiment_rank.get(x, len(sentiment_priority)))
        sentiment_map[post_id] = ordered[0]

        best_flag = 'Unknown'
        best_score = -1
        for flag in group['established_creator_flag']:
            score = creator_rank.get(flag, -1)
            if score > best_score:
                best_flag = flag
                best_score = score
        creator_map[post_id] = best_flag

    snippets_df = snippets_df.copy()
    snippets_df['post_overall_driver'] = snippets_df['post_id'].map(driver_map)
    snippets_df['post_overall_sentiment'] = snippets_df['post_id'].map(sentiment_map)
    snippets_df['post_established_flag'] = snippets_df['post_id'].map(creator_map)
    return snippets_df


def build_examples(snippets_df: pd.DataFrame, posts: list[dict]) -> pd.DataFrame:
    post_map = {p['post_id']: p for p in posts}
    keep = select_example_posts(posts)
    rows = []
    for row in snippets_df.itertuples():
        if row.post_id not in keep:
            continue
        post = post_map[row.post_id]
        rows.append(
            {
                'post_id': row.post_id,
                'platform': post['platform'],
                'source_url': post['url'],
                'raw_verbatim': post['text'],
                'snippet_id': row.snippet_id,
                'verbatim_excerpt': row.verbatim_excerpt,
                'agent_interpretation': row.agent_interpretation,
                'driver_topic': row.driver_topic,
                'sentiment': row.sentiment,
                'app_signal': row.app_signal,
                'post_overall_driver': row.post_overall_driver,
                'post_overall_sentiment': row.post_overall_sentiment,
                'post_established_flag': row.post_established_flag,
            }
        )
    return pd.DataFrame(rows)


def write_workbook(
    posts_df: pd.DataFrame,
    snippets_df: pd.DataFrame,
    examples_df: pd.DataFrame,
    output_path: Path,
) -> None:
    readme_df = pd.DataFrame(
        [
            {'Item': 'Purpose', 'Details': 'Snippets_50 pairs verbatim excerpts with interpretations, rationales, and driver coding.'},
            {'Item': 'Audit Trail', 'Details': 'Posts_Audit links every snippet back to the captured verbatim + URL.'},
            {'Item': 'Topics', 'Details': ', '.join(sorted(snippets_df['driver_topic'].unique()))},
            {'Item': 'Sentiment Scale', 'Details': 'Positive / Negative / Neutral / Mixed (per snippet).'},
            {'Item': 'Equity Measures', 'Details': 'Connection / Discovery / Expression / Entertainment / Creativity / Content Quality signals.'},
            {'Item': 'Reviewer Workflow', 'Details': 'Use dropdowns to adjust drivers, sentiments, signals, creator flags, and review action.'},
            {'Item': 'Certification', 'Details': 'All snippets trace to manual_sample_posts.json; no fabricated text.'},
        ]
    )

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        readme_df.to_excel(writer, sheet_name='README', index=False)
        posts_df.to_excel(writer, sheet_name='Posts_Audit', index=False)
        snippets_df.to_excel(writer, sheet_name='Snippets_50', index=False)
        examples_df.to_excel(writer, sheet_name='Examples10', index=False)

    wb = load_workbook(output_path)
    lists_ws = wb.create_sheet('Lists')
    lists_ws.sheet_state = 'hidden'

    sentiment_priority = ['Negative', 'Mixed', 'Neutral', 'Positive']
    reference_lists = [
        ('driver_topic', sorted(snippets_df['driver_topic'].unique())),
        ('sentiment', sorted(snippets_df['sentiment'].unique())),
        ('app_signal', sorted(snippets_df['app_signal'].unique())),
        ('age_inference', sorted(snippets_df['age_inference'].unique())),
        ('established_creator_flag', ['Unknown', 'Possible', 'Confirmed']),
        ('connection_signal', ['None', 'Possible', 'Strong']),
        ('discovery_signal', ['None', 'Possible', 'Strong']),
        ('expression_signal', ['None', 'Possible', 'Strong']),
        ('entertainment_signal', ['None', 'Possible', 'Strong']),
        ('creativity_signal', ['None', 'Possible', 'Strong']),
        ('content_quality_signal', ['None', 'Possible', 'Strong']),
        ('post_overall_driver', sorted(snippets_df['driver_topic'].unique())),
        ('post_overall_sentiment', sentiment_priority),
        ('post_established_flag', ['Unknown', 'Possible', 'Confirmed']),
        ('review_action', ['Pending', 'Approved', 'Recode']),
    ]

    column_lookup: dict[str, int] = {}
    for col_idx, (name, values) in enumerate(reference_lists, start=1):
        lists_ws.cell(row=1, column=col_idx, value=name)
        for row_idx, value in enumerate(values, start=2):
            lists_ws.cell(row=row_idx, column=col_idx, value=value)
        column_lookup[name] = col_idx

    snippets_ws = wb['Snippets_50']
    posts_ws = wb['Posts_Audit']
    examples_ws = wb['Examples10']
    readme_ws = wb['README']

    for ws in (snippets_ws, posts_ws, examples_ws, readme_ws):
        ws.freeze_panes = 'A2'

    col_widths = {
        'Snippets_50': {
            'A': 12, 'B': 14, 'C': 55, 'D': 60, 'E': 48,
            'F': 22, 'G': 16, 'H': 20, 'I': 18, 'J': 18,
            'K': 16, 'L': 16, 'M': 16, 'N': 16, 'O': 16,
            'P': 16, 'Q': 22, 'R': 22, 'S': 18, 'T': 16,
            'U': 16, 'V': 45,
        },
        'Posts_Audit': {'A': 12, 'B': 14, 'C': 55, 'D': 80},
        'Examples10': {'A': 12, 'B': 14, 'C': 50, 'D': 80, 'E': 16, 'F': 60, 'G': 20, 'H': 16, 'I': 20, 'J': 20, 'K': 18, 'L': 18, 'M': 18},
        'README': {'A': 22, 'B': 90},
    }
    for sheet_name, widths in col_widths.items():
        ws = wb[sheet_name]
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    wrap = Alignment(wrap_text=True, vertical='top')
    for ws in (snippets_ws, posts_ws, examples_ws, readme_ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    max_row = snippets_ws.max_row
    start_ord = ord('F')
    for idx, (name, values) in enumerate(reference_lists):
        column_letter = chr(start_ord + idx)
        col_idx = column_lookup[name]
        formula = f"=Lists!${chr(64 + col_idx)}$2:${chr(64 + col_idx)}${len(values) + 1}"
        dv = DataValidation(type='list', formula1=formula, allow_blank=False)
        snippets_ws.add_data_validation(dv)
        dv.add(f'{column_letter}2:{column_letter}{max_row}')

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Generate the IG preference master workbook.')
    parser.add_argument('--posts', type=Path, default=DEFAULT_POSTS, help='Path to manual_sample_posts.json')
    parser.add_argument('--snippets', type=Path, default=DEFAULT_SNIPPETS, help='Path to manual snippet YAML')
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT, help='Destination xlsx path')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    posts_df, raw_posts = load_posts(args.posts)
    snippets_df = load_snippets(args.snippets)
    enriched_snippets = add_verbatims_and_signals(snippets_df, raw_posts)
    examples_df = build_examples(enriched_snippets, raw_posts)
    write_workbook(posts_df, enriched_snippets, examples_df, args.output)
    print(f"Master workbook rebuilt at {args.output}")


if __name__ == '__main__':
    main()
